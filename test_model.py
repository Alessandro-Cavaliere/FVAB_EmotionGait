import csv
import cv2
import openpyxl
import os
import mediapipe as mp
import pandas as pd
import numpy as np
import argparse
import shutil
from sklearn.preprocessing import LabelEncoder
from keras.models import load_model
from sklearn.metrics import classification_report, confusion_matrix

"""
    Funzione **calcola_media_colonne** per calcolare la media delle colonne in un file Excel.
    Prende in input il percorso del file Excel.
    Restituisce una lista di tuple, ognuna contenente il nome del video e l'etichetta corrispondente.
"""
def calcola_media_colonne(file_excel):
    wb = openpyxl.load_workbook(file_excel)
    foglio = wb.worksheets[0]

    media_colonne = []
    etichette = []
    nomi_video = []

    for riga in foglio.iter_rows(min_row=2, max_row=2, min_col=2):
        for cella in riga:
            if isinstance(cella.value, str):
                if cella.value.startswith("The person in the video appears"):
                    etichetta = "-" + cella.value.split("-")[-1].strip()
                else:
                    etichetta = cella.value.strip()
                etichette.append(etichetta)
            else:
                etichette.append('')

    for cella in foglio[1][1:]:
        if isinstance(cella.value, str):
            nomi_video.append(cella.value)
        else:
            nomi_video.append('')

    for colonna in foglio.iter_cols(min_row=3, min_col=2):
        valori_colonna = [cella.value for cella in colonna if isinstance(cella.value, (int, float))]
        if valori_colonna:
            media = sum(valori_colonna) / len(valori_colonna)
        else:
            media = 0
        media_colonne.append(media)

    risultato = []
    ultimo_nome_video = ''

    for i in range(len(media_colonne)):
        sub_array = media_colonne[i:i + 4]
        if sub_array:
            indice_massimo = sub_array.index(max(sub_array))
            massimo = max(sub_array)
            etichetta = etichette[i + indice_massimo] if i + indice_massimo < len(etichette) else ''
            nome_video = nomi_video[i] if i < len(nomi_video) else ''
            if nome_video != ultimo_nome_video:
                risultato.append((nome_video, etichetta))
                ultimo_nome_video = nome_video
    return risultato

"""
    Funzione **apply_landmarks** per applicare i landmarks a un frame.
    Prende in input il frame, l'oggetto mediapipe, la larghezza e l'altezza desiderate.
    Restituisce il frame con i landmarks applicati.
"""
def apply_landmarks(frame, mp_holistic, target_width, target_height):
    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    frame_rgb = cv2.GaussianBlur(frame_rgb, (5, 5), 0)
    frame_rgb = cv2.resize(frame_rgb, (target_width, target_height))

    results = mp_holistic.process(frame_rgb)
    frame_with_landmarks = frame_rgb.copy()
    mp.solutions.drawing_utils.draw_landmarks(
        frame_with_landmarks, results.pose_landmarks, mp.solutions.holistic.POSE_CONNECTIONS)
    mp.solutions.drawing_utils.draw_landmarks(
        frame_with_landmarks, results.face_landmarks)

    return frame_with_landmarks

"""
    Funzione **process_video_folder** per processare una cartella di video.
    Prende in input il percorso della cartella di input, il percorso della cartella di output, la larghezza e l'altezza desiderate.
    Restituisce una lista di tuple, ognuna contenente il nome del video e la lista dei frame con i landmarks applicati.
"""
def process_video_folder(input_folder, output_folder, target_width, target_height):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    mp_holistic = mp.solutions.holistic.Holistic()
    processed_videos = []

    for filename in os.listdir(input_folder):
        if filename.lower().endswith(('.avi', '.mp4', '.mkv', '.mov')):
            video_path = os.path.join(input_folder, filename)
            video_name = os.path.splitext(filename)[0]
            output_subfolder = os.path.join(output_folder, video_name)
            index = 1
            while os.path.exists(output_subfolder):
                output_subfolder = os.path.join(output_folder, f"{video_name}_{index}")
                index += 1
            os.makedirs(output_subfolder)

            cap = cv2.VideoCapture(video_path)
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_with_landmarks_list = []
            frame_count = 0
            start_time = 0
            end_time = start_time + 1

            while True:
                ret, frame = cap.read()
                if not ret:
                    break

                current_time = frame_count / fps

                if current_time >= start_time and current_time < end_time:
                    frame_with_landmarks = apply_landmarks(frame, mp_holistic, target_width, target_height)
                    frame_with_landmarks_list.append(frame_with_landmarks)
                    output_path = os.path.join(output_subfolder, f"frame{frame_count}.jpg")
                    cv2.imwrite(output_path, frame_with_landmarks)

                frame_count += 1

                if current_time >= end_time:
                    break

            cap.release()

            processed_videos.append((video_name, frame_with_landmarks_list))

    mp_holistic.close()

    return processed_videos

"""
    Funzione **test_lstm** per testare un modello LSTM su un set di dati.
    Prende in input il percorso del file CSV e una lista di nomi di cartelle.
    Non restituisce nulla, ma stampa l'accuratezza del modello sul set di test e le metriche di classificazione.
"""
def test_lstm(file_csv, nomi_cartelle,model):
    df = pd.read_csv(file_csv)
    emotions = {'-Happy': 0, '-Sad': 1, '-Angry': 2, '-Neutral': 3}

    X = []
    y = []

    # Iterazione sui nomi delle cartelle
    for folder_name in nomi_cartelle:
        video_name = " " + folder_name  # Nome della cartella video

        if video_name in df['Nome Video'].values:
            emotion = df.loc[df['Nome Video'] == video_name, 'Emozione'].values[0]
            emotion_label = emotions[emotion]

            frame_array = []  # Array per salvare i frame

            folder_path = os.path.abspath("outputframeTest\\" + folder_name)
            framesInVideoDir = os.listdir(folder_path)
            framesInVideoDirSorted = sorted(framesInVideoDir, key=lambda x: int(''.join(filter(str.isdigit, x))))

            for frame_file in framesInVideoDirSorted:
                frame_path = os.path.join(folder_path, frame_file)
                frame = cv2.imread(frame_path)
                frame = cv2.resize(frame, (90, 160))  # Riduci la dimensione dell'immagine
                frame_array.append(frame)

            X.append(frame_array)  # Aggiungi l'array dei frame con landmarks
            y.append(emotion_label)

    print(y)
    # Codifica delle etichette
    le = LabelEncoder()
    y = le.fit_transform(y)

    X = np.array(X)
    y = np.array(y)

    # Load the pre-trained model
    testModel = load_model(model)

    # Evaluate the model on the test data
    test_loss, test_accuracy = testModel.evaluate(X, y)
    print("Accuratezza sul set di test: {:.2f}%".format(test_accuracy * 100))

    # Make predictions on the test data
    y_pred = testModel.predict(X)
    y_pred_classes = np.argmax(y_pred, axis=1)

    # Print classification report
    print(classification_report(y, y_pred_classes, target_names=emotions.keys()))

    # Print confusion matrix
    print(confusion_matrix(y, y_pred_classes))

"""
    Funzione **get_folder_names** per ottenere i nomi delle sottocartelle in una cartella.
    Prende in input il percorso della cartella.
    Restituisce una lista di nomi di sottocartelle.
"""
def get_folder_names(output_folder):
    if os.path.isdir(output_folder):
        nomi_cartelle = [nome for nome in os.listdir(output_folder) if os.path.isdir(os.path.join(output_folder, nome))]
        return nomi_cartelle
    else:
        print('Il percorso specificato non corrisponde a una cartella.')
        return []

def main():
    # Crea un parser per gli argomenti da riga di comando
    parser = argparse.ArgumentParser(
        description='Parser per definire gli argomenti da riga di comando (il file .xlsx e la cartella dei video).')
    parser.add_argument('--file', type=str, help='Il percorso del file Excel da processare.')
    parser.add_argument('--video-folder', type=str, help='La cartella contenente i video da elaborare.')
    parser.add_argument('--model', type=str, help='Modello da voler testare.')

    # Analizza gli argomenti da riga di comando
    args = parser.parse_args()

    # Se l'utente ha fornito un percorso di file, usalo. Altrimenti, usa il percorso di default.
    if args.file:
        file_excel = args.file
    else:
        print(
            "\nStai per utilizzare il percorso di default per il file Excel: 'datasets_xlsx/responses_test.xlsx'")
        print("Vuoi continuare? [y/n]")
        response = input()
        if response.lower() != 'y':
            print(
                "\nPer specificare un percorso di file diverso, esegui lo script con l'opzione \033[91m--file\033[0m, come segue:")
            print(
                "      \033[91mpython\033[0m test_model.py \033[91m--file\033[0m /percorso/del/tuo/file.xlsx\n")
            exit()
        file_excel = 'datasets_xlsx/responses_test.xlsx'

    risultato = calcola_media_colonne(file_excel)

    # Controllo se è stata fornita una cartella dei video per il test diversa tramite l'argomento --video-folder
    if args.video_folder:
        input_folder = args.video_folder
    elif os.path.exists("./videosTest"):
        print("\nStai per utilizzare il percorso di default per la cartella dei video di test: './videosTest'")
        print("Vuoi continuare? [y/n]")
        response = input()
        if response.lower() != 'y':
            print("\nPer specificare un percorso della cartella dei video di test diverso, esegui lo script con l'opzione \033[91m--video-folder\033[0m, come segue:")
            print("      \033[91mpython\033[0m test_model.py \033[91m--video-folder\033[0m /percorso/della/tua/cartellaDeiVideoDiTest\n")
            exit()
        else:
            input_folder = './videosTest'
            print("Continuazione dello script in corso...")
    else:
        print(
            "\nÈ necessario specificare un percorso della cartella dei video diverso, esegui lo script con l'opzione \033[91m--video-folder\033[0m, come segue:")
        print("      \033[91mpython\033[0m test_model.py \033[91m--video-folder\033[0m /percorso/della/tua/cartellaDeiVideoDiTest\n")
        print("Oppure importa una cartella 'videosTest' con all'interno i video di test di tuo interesse nella workspace corrente.")
        exit()

    # Controllo se è stata fornito un modello da testare tramite l'argomento --model
    if args.model:
        model=args.model
    elif os.path.exists("emotion_lstm_model.h5"):
        print("\nStai per utilizzare il modello corrente in locale: 'emotion_lstm_model.h5'")
        print("Vuoi continuare? [y/n]")
        response = input()
        if response.lower() != 'y':
            print(
                "\nPer specificare un modello da testare , esegui lo script con l'opzione \033[91m--model\033[0m, come segue:")
            print(
                "      \033[91mpython\033[0m test_model.py \033[91m--model\033[0m /percorso/del/tuo/modello\n")
            exit()
        else:
            model = 'emotion_lstm_model.h5'
            print("Continuazione dello script in corso...")
    else:
        print("\nÈ necessario specificare il percorso del modello da testare, esegui lo script con l'opzione \033[91m--model\033[0m, come segue:")
        print("      \033[91mpython\033[0m test_model.py \033[91m--model\033[0m /percorso/del/tuo/modello\n")
        print("Oppure importa un modello con nome 'emotion_lstm_model.h5' nella workspace corrente.")
        exit()

    file_csv = './dativideoTest.csv'
    with open(file_csv, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Nome Video', 'Emozione'])
        for nome_video, etichetta in risultato:
            if nome_video.startswith("VID_RGB_000"):
                nome_video = nome_video.replace("VID_RGB_000", "").strip()
            writer.writerow([nome_video, etichetta])
            print("Nome video:", nome_video)
            print("Emozione:", etichetta)
            print()

    output_folder = './outputframeTest'
    target_width = 180
    target_height = 320

    if os.path.exists("./outputframeTest") | os.path.exists("emotion_lstm_model.h5"):
        print("\nSiccome esistono rimasugli di vecchie esecuzioni dello script, è necessario cancellarle per assicurarsi che l'esecuzione avvenga senza intoppi.")
        print("Assicurati che questi file sensibili non ti siano utili (\033[91m./outputframeTest\033[0m - \033[91memotion_lstm_model.h5\033[0m)")
        print("Procedo alla cancellazione? [y/n]")
        response = input()
        if response.lower() != 'y':
            print("\nÈ necessario che questi file vengano spostati da questa workspace per il corretto funzionamento.")
            print("Riavvia lo script una volta che questi file siano stati spostati (o cancellati) correttamente.")
            exit()
        else:
            # Cancella la cartella 'outputframe' e tutto il suo contenuto nel caso siano rimasti rimasugli da vecchie esecuzioni
            if os.path.exists("./outputframeTest"):
                shutil.rmtree('./outputframeTest')
            if os.path.exists("emotion_lstm_model.h5"):
                os.remove('emotion_lstm_model.h5')
            print("\nCancellazione dei file avvenuta con successo.")
            print("Esecuzione dello script in corso...\n\n")

    process_video_folder(input_folder, output_folder, target_width, target_height)
    nomi_cartelle = get_folder_names(output_folder)
    test_lstm(file_csv, nomi_cartelle,model)


if __name__ == '__main__':
    main()