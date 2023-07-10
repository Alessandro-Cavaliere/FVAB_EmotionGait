# Importazione di tutti i vari moduli utilizzati per questo file.
import csv
import cv2
import openpyxl
import os
import numpy as np
import mediapipe as mp
import pandas as pd
import argparse
import shutil
import matplotlib.pyplot as plt
from keras.models import Sequential
from keras.regularizers import l2
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import f1_score, confusion_matrix
from keras.layers import TimeDistributed, Conv2D, MaxPooling2D, Flatten, LSTM, Dense
from keras.callbacks import EarlyStopping, TensorBoard
from sklearn.utils import compute_sample_weight
from imblearn.over_sampling import RandomOverSampler
from keras.optimizers import Adam

"""
    Funzione **calcola_media_colonne** per calcolare la media delle colonne in un file Excel.
    Prende in input il percorso del file Excel.
    Restituisce una lista di tuple, ognuna contenente il nome del video e l'etichetta corrispondente.
"""


def calcola_media_colonne(file_excel):
    wb = openpyxl.load_workbook(file_excel)
    foglio = wb.worksheets[0]

    media_colonne = []
    etichette = []
    nomi_video = []

    for riga in foglio.iter_rows(min_row=2, max_row=2, min_col=2):
        for cella in riga:
            if isinstance(cella.value, str):
                if cella.value.startswith("The person in the video appears"):
                    etichetta = "-" + cella.value.split("-")[-1].strip()
                else:
                    etichetta = cella.value.strip()
                etichette.append(etichetta)
            else:
                etichette.append('')

    for cella in foglio[1][1:]:
        if isinstance(cella.value, str):
            nomi_video.append(cella.value)
        else:
            nomi_video.append('')

    for colonna in foglio.iter_cols(min_row=3, min_col=2):
        valori_colonna = [cella.value for cella in colonna if isinstance(cella.value, (int, float))]
        if valori_colonna:
            media = sum(valori_colonna) / len(valori_colonna)
        else:
            media = 0
        media_colonne.append(media)

    risultato = []
    ultimo_nome_video = ''

    for i in range(len(media_colonne)):
        sub_array = media_colonne[i:i + 4]
        if sub_array:
            indice_massimo = sub_array.index(max(sub_array))
            massimo = max(sub_array)
            etichetta = etichette[i + indice_massimo] if i + indice_massimo < len(etichette) else ''
            nome_video = nomi_video[i] if i < len(nomi_video) else ''
            if nome_video != ultimo_nome_video:
                risultato.append((nome_video, etichetta))
                ultimo_nome_video = nome_video
    return risultato


"""
   Funzione **apply_landmarks** per applicare i landmarks a un frame.
   Prende in input il frame, l'oggetto mediapipe, la larghezza e l'altezza desiderate.
   Restituisce il frame con i landmarks applicati.
"""


def apply_landmarks(frame, mp_holistic, target_width, target_height):
    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    frame_rgb = cv2.GaussianBlur(frame_rgb, (5, 5), 0)
    frame_rgb = cv2.resize(frame_rgb, (target_width, target_height))

    results = mp_holistic.process(frame_rgb)
    frame_with_landmarks = frame_rgb.copy()
    mp.solutions.drawing_utils.draw_landmarks(
        frame_with_landmarks, results.pose_landmarks, mp.solutions.holistic.POSE_CONNECTIONS)
    mp.solutions.drawing_utils.draw_landmarks(
        frame_with_landmarks, results.face_landmarks)

    return frame_with_landmarks


"""
    Funzione **process_video_folder** per processare una cartella di video.
    Prende in input il percorso della cartella di input, il percorso della cartella di output, la larghezza e l'altezza desiderate.
    Restituisce una lista di tuple, ognuna contenente il nome del video e la lista dei frame con i landmarks applicati.
"""


def process_video_folder(input_folder, output_folder, target_width, target_height):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    mp_holistic = mp.solutions.holistic.Holistic()
    processed_videos = []

    for filename in os.listdir(input_folder):
        if filename.lower().endswith(('.avi', '.mp4', '.mkv', '.mov')):
            video_path = os.path.join(input_folder, filename)
            video_name = os.path.splitext(filename)[0]
            output_subfolder = os.path.join(output_folder, video_name)
            index = 1
            while os.path.exists(output_subfolder):
                output_subfolder = os.path.join(output_folder, f"{video_name}_{index}")
                index += 1
            os.makedirs(output_subfolder)

            cap = cv2.VideoCapture(video_path)
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_with_landmarks_list = []
            frame_count = 0
            start_time = 0
            end_time = start_time + 1

            while True:
                ret, frame = cap.read()
                if not ret:
                    break

                current_time = frame_count / fps

                if current_time >= start_time and current_time < end_time:
                    frame_with_landmarks = apply_landmarks(frame, mp_holistic, target_width, target_height)
                    frame_with_landmarks_list.append(frame_with_landmarks)
                    output_path = os.path.join(output_subfolder, f"frame{frame_count}.jpg")
                    cv2.imwrite(output_path, frame_with_landmarks)

                frame_count += 1

                if current_time >= end_time:
                    break

            cap.release()

            processed_videos.append((video_name, frame_with_landmarks_list))

    mp_holistic.close()

    return processed_videos


"""
    Funzione **train_lstm** per addestrare un modello LSTM.
    Prende in input il percorso del file CSV e una lista di nomi di cartelle.
    Non restituisce nulla, ma salva il modello addestrato in un file .h5.
"""


def train_lstm(file_csv, nomi_cartelle):
    df = pd.read_csv(file_csv)
    emotions = {'-Happy': 0, '-Sad': 1, '-Angry': 2, '-Neutral': 3}

    X = []
    y = []

    # Iterazione sui nomi delle cartelle
    for folder_name in nomi_cartelle:
        video_name = " " + folder_name  # Nome della cartella video

        if video_name in df['Nome Video'].values:
            emotion = df.loc[df['Nome Video'] == video_name, 'Emozione'].values[0]
            emotion_label = emotions[emotion]

            frame_array = []  # Array per salvare i frame

            folder_path = os.path.abspath("outputframe\\" + folder_name)
            framesInVideoDir = os.listdir(folder_path)
            framesInVideoDirSorted = sorted(framesInVideoDir, key=lambda x: int(''.join(filter(str.isdigit, x))))

            for frame_file in framesInVideoDirSorted:
                frame_path = os.path.join(folder_path, frame_file)
                frame = cv2.imread(frame_path)
                frame = cv2.resize(frame, (90, 160))  # Riduci la dimensione dell'immagine
                frame_array.append(frame)

            X.append(frame_array)  # Aggiungi l'array dei frame con landmarks
            y.append(emotion_label)

    print(y)
    # Codifica le etichette delle emozioni in numeri
    le = LabelEncoder()
    y = le.fit_transform(y)

    # Trasforma le liste X e y in array numpy
    X = np.array(X)
    y = np.array(y)

    # Ottieni il numero di timesteps e la forma di input per il modello
    timesteps = X.shape[1]
    print("TIMESTEP:\n")
    print(timesteps)
    input_shape = X.shape[2:]
    print("INPUP:\n")
    print(input_shape)

    # Applica l'oversampling per bilanciare le classi
    ros = RandomOverSampler(random_state=42)
    X_res, y_res = ros.fit_resample(X.reshape(len(X), -1), y)
    X_res = X_res.reshape(-1, timesteps, input_shape[0], input_shape[1], input_shape[2])

    # Suddividi i dati in set di addestramento e di test
    X_train, X_test, y_train, y_test = train_test_split(X_res, y_res, test_size=0.2, random_state=42)

    train_label_distribution = np.bincount(y_train)
    train_labels = np.unique(y_train)

    # Calcola la distribuzione delle etichette nel set di validazione
    val_label_distribution = np.bincount(y_test)
    val_labels = np.unique(y_test)

    # Plot della distribuzione delle etichette nel set di addestramento
    plt.figure(figsize=(8, 5))
    plt.bar(train_labels, train_label_distribution)
    plt.xlabel('Etichette')
    plt.ylabel('Numero di campioni')
    plt.title('Distribuzione delle etichette nel set di addestramento')
    plt.show()

    # Plot della distribuzione delle etichette nel set di validazione
    plt.figure(figsize=(8, 5))
    plt.bar(val_labels, val_label_distribution)
    plt.xlabel('Etichette')
    plt.ylabel('Numero di campioni')
    plt.title('Distribuzione delle etichette nel set di validazione')
    plt.show()


    # Calcola i pesi delle classi per bilanciare ulteriormente le classi durante l'addestramento
    sample_weights = compute_sample_weight(class_weight='balanced', y=y_train)

    """
        Crea il modello LSTM:
        1. Aggiungi un livello di convoluzione 2D con 16 filtri, una dimensione del kernel di 3x3 e una funzione di attivazione ReLU.
           Il livello TimeDistributed permette di applicare la convoluzione a ciascuno dei timesteps indipendentemente.
        2. Aggiungi un livello di pooling per ridurre la dimensione spaziale dell'output del livello precedente.
        3. Aggiungi un altro livello di convoluzione 2D, questa volta con 32 filtri.
        4. Aggiungi un altro livello di pooling.
        5. Appiattisci l'output del livello precedente per poterlo passare a un livello LSTM.
        6. Aggiungi un livello LSTM con 32 unità. Questo livello analizza le sequenze di output del livello precedente nel tempo.
        7. Aggiungi un livello completamente connesso (Dense) con 32 unità e una funzione di attivazione ReLU.
           Il regularizzatore l2 aiuta a prevenire l'overfitting penalizzando i pesi grandi.
        8. Aggiungi un livello di output completamente connesso con un numero di unità pari al numero di classi uniche in y.
           La funzione di attivazione softmax garantisce che l'output possa essere interpretato come probabilità per ciascuna classe.
    """
    model = Sequential()
    model.add(TimeDistributed(Conv2D(16, (3, 3), activation="relu"), input_shape=(timesteps,) + input_shape))
    model.add(TimeDistributed(MaxPooling2D((2, 2))))
    model.add(TimeDistributed(Conv2D(32, (3, 3), activation="relu")))
    model.add(TimeDistributed(MaxPooling2D((2, 2))))
    model.add(TimeDistributed(Flatten()))
    model.add(LSTM(32, return_sequences=False))
    model.add(Dense(32, activation='relu', kernel_regularizer=l2(0.001)))
    model.add(Dense(len(np.unique(y)), activation="softmax"))

    # Stampa un sommario del modello
    model.summary()

    # Compila il modello con un learning rate specifico
    opt = Adam(learning_rate=0.00007)
    model.compile(loss='sparse_categorical_crossentropy', optimizer=opt, metrics=['accuracy'])

    # Crea i callback per l'addestramento
    early_stopping = EarlyStopping(monitor='val_accuracy', patience=20)
    # reduce_lr = ReduceLROnPlateau(monitor='val_loss', factor=0.2, patience=5, min_lr=0.00001)  # Nuovo callback
    tensorboard_callback = TensorBoard(log_dir='./logs', histogram_freq=1)

    # Addestra il modello --> Inizio addestramento
    print("Inizio addestramento del modello LSTM...")
    history = model.fit(X_train, y_train, batch_size=32, epochs=150, validation_data=(X_test, y_test),
                        callbacks=[early_stopping, tensorboard_callback],
                        sample_weight=sample_weights)

    # Stampa l'accuratezza del modello sul set di addestramento
    accuracy = history.history['accuracy'][-1]
    print("Accuratezza del modello LSTM: {:.2f}%".format(accuracy * 100))

    # Valuta il modello sul set di test e stampa l'accuratezza
    test_loss, test_accuracy = model.evaluate(X_test, y_test)
    print("Accuratezza sul set di test: {:.2f}%".format(test_accuracy * 100))

    # Salvataggio del modello addestrato
    model.save('emotion_lstm_model.h5')
    print("Modello LSTM addestrato salvato con successo.\n")

    y_pred = model.predict(X_test)
    y_pred_classes = np.argmax(y_pred, axis=1)

    f1score = f1_score(y_test, y_pred_classes, average='weighted')
    print("F1 Score: " + str(f1score))

    cm = confusion_matrix(y_test, y_pred_classes)
    class_accuracy = cm.diagonal() / cm.sum(axis=1)

    for i, accuracy in enumerate(class_accuracy):
        emotion = le.inverse_transform([i])[0]
        if emotion == 0:
            print(f"Accuracy for Happy: {accuracy}")
        elif emotion == 1:
            print(f"Accuracy for Sad: {accuracy}")
        elif emotion == 2:
            print(f"Accuracy for Angry: {accuracy}")
        elif emotion == 3:
            print(f"Accuracy for Neutral: {accuracy}")
        else:
            print("Invalid emotion")


"""
    Funzione **get_folder_names** per ottenere i nomi delle sottocartelle in una cartella.
    Prende in input il percorso della cartella.
    Restituisce una lista di nomi di sottocartelle.
"""


def get_folder_names(output_folder):
    if os.path.isdir(output_folder):
        nomi_cartelle = [nome for nome in os.listdir(output_folder) if os.path.isdir(os.path.join(output_folder, nome))]
        return nomi_cartelle
    else:
        print('Il percorso specificato non corrisponde a una cartella.')
        return []


def main():
    # Crea un parser per gli argomenti da riga di comando
    parser = argparse.ArgumentParser(
        description='Parser per definire gli argomenti da riga di comando (il file .xlsx e la cartella dei video).')
    parser.add_argument('--file', type=str, help='Il percorso del file Excel da processare.')
    parser.add_argument('--video-folder', type=str, help='La cartella contenente i video da elaborare.')

    # Analizza gli argomenti da riga di comando
    args = parser.parse_args()

    # Se l'utente ha fornito un percorso di file, usalo. Altrimenti, usa il percorso di default.
    if args.file:
        file_excel = args.file
    else:
        print(
            "\nStai per utilizzare il percorso di default per il file Excel: 'datasets_xlsx/Responses_train+validation.xlsx'")
        print("Vuoi continuare? [y/n]")
        response = input()
        if response.lower() != 'y':
            print(
                "\nPer specificare un percorso di file diverso, esegui lo script con l'opzione \033[91m--file\033[0m, come segue:")
            print(
                "      \033[91mpython\033[0m test_validation_model.py \033[91m--file\033[0m /percorso/del/tuo/file.xlsx\n")
            exit()
        file_excel = 'datasets_xlsx/Responses_train+validation.xlsx'

    risultato = calcola_media_colonne(file_excel)

    # Controllo se è stata fornita una cartella dei video diversa tramite l'argomento --video-folder
    if args.video_folder:
        input_folder = args.video_folder
    elif os.path.exists("./videos"):
        print("\nStai per utilizzare il percorso di default per la cartella dei video: './videos'")
        print("Vuoi continuare? [y/n]")
        response = input()
        if response.lower() != 'y':
            print(
                "\nPer specificare un percorso della cartella dei video diverso, esegui lo script con l'opzione \033[91m--video-folder\033[0m, come segue:")
            print(
                "      \033[91mpython\033[0m test_validation_model.py \033[91m--video-folder\033[0m /percorso/della/tua/cartellaDeiVideo\n")
            exit()
        else:
            input_folder = './videos'
            print("Continuazione dello script in corso...")
    else:
        print(
            "\nÈ necessario specificare un percorso della cartella dei video diverso, esegui lo script con l'opzione \033[91m--video-folder\033[0m, come segue:")
        print(
            "      \033[91mpython\033[0m test_validation_model.py \033[91m--video-folder\033[0m /percorso/della/tua/cartellaDeiVideo\n")
        print("Oppure importa una cartella 'videos' con all'interno i video di tuo interesse nella workspace corrente.")
        exit()

    file_csv = './dativideo.csv'
    with open(file_csv, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Nome Video', 'Emozione'])
        for nome_video, etichetta in risultato:
            if nome_video.startswith("VID_RGB_000"):
                nome_video = nome_video.replace("VID_RGB_000", "").strip()
            writer.writerow([nome_video, etichetta])
            print("Nome video:", nome_video)
            print("Emozione:", etichetta)
            print()

    print("Dati salvati correttamente nel file CSV:", file_csv)

    output_folder = './outputframe'
    target_width = 180
    target_height = 320

    if os.path.exists("./outputframe") | os.path.exists("./logs") | os.path.exists("emotion_lstm_model.h5"):
        print(
            "\nSiccome esistono rimasugli di vecchie esecuzioni dello script, è necessario cancellarle per assicurarsi che l'esecuzione avvenga senza intoppi.")
        print(
            "Assicurati che questi file sensibili non ti siano utili (\033[91m./outputframe\033[0m - \033[91m./logs\033[0m - \033[91memotion_lstm_model.h5\033[0m)")
        print("Procedo alla cancellazione? [y/n]")
        response = input()
        if response.lower() != 'y':
            print("\nÈ necessario che questi file vengano spostati da questa workspace per il corretto funzionamento.")
            print("Riavvia lo script una volta che questi file siano stati spostati (o cancellati) correttamente.")
            exit()
        else:
            # Cancella la cartella 'outputframe' e tutto il suo contenuto nel caso siano rimasti rimasugli da vecchie esecuzioni
            if os.path.exists("./outputframe"):
                print("no")
                #shutil.rmtree('./outputframe')
            if os.path.exists("./logs"):
                shutil.rmtree('./logs')
            if os.path.exists("emotion_lstm_model.h5"):
                os.remove('emotion_lstm_model.h5')
            print("\nCancellazione dei file avvenuta con successo.")
            print("Esecuzione dello script in corso...\n\n")

    if not os.path.exists("./outputframe"):
        process_video_folder(input_folder, output_folder, target_width, target_height)
    nomi_cartelle = get_folder_names(output_folder)
    train_lstm(file_csv, nomi_cartelle)


if __name__ == '__main__':
    main()
